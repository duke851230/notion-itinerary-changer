from __future__ import annotations

from typing import TYPE_CHECKING, Any, List, Optional, Union

if TYPE_CHECKING:
    pass

from constant.excel import ColorMap
from helpers.excel.utils import (
    find_next_column_letter,
    get_default_border,
    get_type_color,
    get_type_display_text,
)
from helpers.general import exchange_place_with_key_and_value
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


def initialize_sheet(sheet: Worksheet, merge_row_num: int) -> None:
    """Initialize excel sheet.

    這邊 defaultRowHeight 會依據一個時間段內要顯示的行數（merge_row_num）來設定行高。

    :param sheet: current sheet
    :param merge_row_num: how many timeline item need to merge into one period

    :return: None
    """
    sheet.sheet_format.defaultColWidth = 20
    sheet.sheet_format.defaultRowHeight = 60 // merge_row_num
    sheet.row_dimensions[1].height = 35


def set_timeline_in_sheet(
    sheet: Worksheet, column: str, timeline_data: List[dict], merge_row_num: int
) -> None:
    """Set schedule's timeline in the excel sheet.

    - 因時間軸分成「人看的」與「電腦看的」兩種。「電腦看的」為較小粒度的時間（例如 5 或 10 分鐘）；「人看的」則像是 30 或 60 分鐘。
      我們的 timeline_data 就是用來給「電腦看的」。要顯示給人看時，需將多筆小粒度時間結合在一起，而 merge_row_num 就是用來表示要將幾筆合併。
      例如：「人看的」為 60 分鐘，「電腦看的」為 10 分鐘，則 merge_row_num 就為 6。
    - 如果當迴圈結束時還有幾筆時間沒有被處理到的話，就把剩下的時間合併成一格

    :param sheet: current sheet
    :param column: the column which is used to write timeline
    :param timeline_data: has two dict, first is start timeline, the second is end timeline
    :param merge_row_num: how many timeline item need to merge into one period

    :return: None
    """
    sheet.column_dimensions[column].width = 12

    start_timeline, end_timeline = timeline_data
    start_timeline = exchange_place_with_key_and_value(start_timeline)
    end_timeline = exchange_place_with_key_and_value(end_timeline)

    need_merge_rows: List[dict] = []
    for row_id, end_time in end_timeline.items():
        need_merge_rows.append({"row_id": row_id, "time": end_time})
        if len(need_merge_rows) < merge_row_num:
            continue

        sheet.merge_cells(
            f"{column}{need_merge_rows[0]['row_id']}:{column}{need_merge_rows[-1]['row_id']}"
        )

        c: Union[MergedCell, Any, Cell] = sheet.cell(
            row=need_merge_rows[0]['row_id'],
            column=column_index_from_string(column),
            value=f"{start_timeline[need_merge_rows[0]['row_id']]}~{need_merge_rows[-1]['time']}",
        )
        set_general_format_of_cell(c, font_size=10, fill_color=ColorMap.gray.value)
        need_merge_rows.clear()

    # 如果當迴圈結束時還有幾筆時間沒有被處理到的話，就把剩下的時間合併成一格
    if need_merge_rows:
        sheet.merge_cells(
            f"{column}{need_merge_rows[0]['row_id']}:{column}{need_merge_rows[-1]['row_id']}"
        )
        c: Union[MergedCell, Any, Cell] = sheet.cell(
            row=need_merge_rows[0]['row_id'],
            column=column_index_from_string(column),
            value=f"{start_timeline[need_merge_rows[0]['row_id']]}~{need_merge_rows[-1]['time']}",
        )
        set_general_format_of_cell(c, font_size=10, fill_color=ColorMap.gray.value)
        need_merge_rows.clear()


def insert_activities_to_sheet(
    sheet: Worksheet, activities: List[list], timeline: dict, start_column: str
) -> None:
    """Insert activities and date title to excel.

    :param sheet: current work sheet object
    :param activities: everyday activities
    :param timeline: schedule's end timeline which is used to merge activity's cell
    :param start_column: define where to start writing

    :return: None
    """
    current_column: str = start_column

    for today_activities in activities:
        header_cell: Union[MergedCell, Any, Cell] = sheet.cell(
            row=1,
            column=column_index_from_string(current_column),
            value=today_activities[0]["date"],
        )
        set_general_format_of_cell(header_cell, font_size=16, fill_color=ColorMap.dark_gray.value)

        for activity in today_activities:
            start_time_str: str = activity["start_at"]
            end_time_str: str = activity["end_at"]
            start_row_index: int = timeline[start_time_str] + 1
            end_row_index: int = timeline[end_time_str]

            sheet.merge_cells(f"{current_column}{start_row_index}:{current_column}{end_row_index}")

            try:
                activity_cell: Union[MergedCell, Any, Cell] = sheet.cell(
                    row=start_row_index,
                    column=column_index_from_string(current_column),
                    value=get_type_display_text(
                        used_rows=end_row_index - start_row_index + 1,
                        type_name=activity["type"],
                        activity_name=activity["name"],
                        start_time=start_time_str,
                        end_time=end_time_str,
                        note=activity["note"],
                    ),
                )
            except Exception as e:
                print(f"activity: {activity}")
                raise e

            set_general_format_of_cell(
                activity_cell, font_size=10, fill_color=get_type_color(activity["type"])
            )
            if isinstance(activity_cell, Cell):
                auto_adjust_cell_width(sheet, activity_cell)

        current_column = find_next_column_letter(current_column)


def set_general_format_of_cell(
    cell: Union[MergedCell, Any, Cell], font_size: int, fill_color: Optional[str]
) -> None:
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(name=u'微軟雅黑', size=font_size)
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.border = get_default_border()


def auto_adjust_cell_width(sheet: Worksheet, cell: Cell) -> None:
    MAX_WIDTH: float = 50.0

    value_list: List[str] = str(cell.value).split("\n")
    value_length_list: List[int] = [len(i) for i in value_list]
    if len(value_length_list) == 1:
        value_max_length: float = value_length_list[0] * 1.7
    else:
        value_max_length = max(*value_length_list) * 1.7

    cell_length: float = max(value_max_length, sheet.column_dimensions[cell.column_letter].width)

    if cell_length < MAX_WIDTH:
        sheet.column_dimensions[cell.column_letter].width = cell_length
    else:
        sheet.column_dimensions[cell.column_letter].width = MAX_WIDTH
